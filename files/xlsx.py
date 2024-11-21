"""Gherkin steps related with XSLX files."""
from __future__ import unicode_literals

from aloe import step
from aloe.tools import guess_types
from nose.tools import assert_equal

from aloe_webdriver_extra.util import CAPTURE_STRING
from .util import wait_for_file


@step(
    r'downloaded XLSX file {STRING}(?: in sheet {STRING})? should'
    r' contain:$'.format(
        STRING=CAPTURE_STRING,
    ))
def check_xlsx_file(self, filename, sheetname=None):
    """
    Check that the given data exists on the XLSX file.

    :param self: Object reference to aloe.
    :param filename: Filename of the XLSX file to verify.
    :param sheetname: Name of the sheet to use. If none is specified then the
        first one will be used.

    Example:
        | A18   | sheet1:A18 |
        | value | value      |
    """

    try:
        from openpyxl import load_workbook
    except ImportError:
        assert False, "OpenPyXL is required for analysing XLSX files."

    assert self.table is not None, 'XLSX content not specified'

    workbook = load_workbook(filename=wait_for_file(filename))
    first_sheet = workbook.get_sheet_names()[0]

    for row in guess_types(self.hashes):
        for cell_coords, value in row.items():
            cell_sheetname = None

            if ':' in cell_coords:
                assert sheetname is None, (
                    'Sheet name must be specified only once.'
                )

                cell_sheetname, cell_coords = cell_coords.split(':')
            elif sheetname:
                cell_sheetname = sheetname

            if cell_sheetname:
                sheet = workbook.get_sheet_by_name(cell_sheetname)

                assert sheet is not None, (
                    'Sheet "{sheet_name}" not found. Available sheets: '
                    '{sheets}'
                ).format(
                    sheet_name=cell_sheetname,
                    sheets=', '.join(workbook.get_sheet_names())
                )
            else:
                sheet = first_sheet

            cell_value = sheet.cell(cell_coords).value
            assert_equal(
                cell_value,
                value,
                'Value [{value}] does not match expected value in cell '
                '{cell}'.format(
                    value=cell_value,
                    cell=cell_coords,
                )
            )
